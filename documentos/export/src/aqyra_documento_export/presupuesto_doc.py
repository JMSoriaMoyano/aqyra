# -*- coding: utf-8 -*-
"""Consumidor CONTRACTUAL del rail firmable: el PRESUPUESTO de obra (E6.2, redireccion 2026-07-12).

El entregable con caracter CONTRACTUAL (junto al pliego): presupuesto POR PARTIDAS del banco de
precios + justificacion de la MEDICION (cantidad + criterio + origen + GUIDs, trazada al modelo) +
CUADROS DE PRECIOS n1 (en letra) y n2 (descompuesto) + resumen PEM->PEC. El artefacto AUTORITATIVO es
el `salida-presupuesto` JSON (C5); este consumidor NO re-renderiza a mano lo que ya existe:
- Word  -> **envuelve** `documentos/presupuesto.componer_documento` (el .docx del despacho, ya firmado).
- BC3   -> **conecta** `engines/bc3.emitir_bc3` (FIEBDC-3, formato de LICITACION publica).
- PDF   -> el firmable sellado (mediciones+justificacion, cuadros n1/n2, resumen, manifiesto embebido).
- XLSX  -> mediciones tabular (para administracion/tecnico).
Determinista, sin LLM. La justificacion de medición v0 = cantidad + criterio + origen + GUIDs (lo que el
`salida-presupuesto` ya trae y ancla GOL-PRE-01); el detalle dimensional por objeto = forward (engine).
"""
from __future__ import annotations

from datetime import datetime
from pathlib import Path

from aqyra_documento_comun import formato as F

_TIPO = {"mano_obra": "Mano de obra", "material": "Material", "maquinaria": "Maquinaria"}


def _fecha(descriptor: dict) -> str:
    return str(descriptor.get("sello_tiempo") or "-")


def _dt(descriptor: dict) -> datetime:
    try:
        return datetime.strptime(_fecha(descriptor), "%Y-%m-%d")
    except ValueError:
        return datetime(1970, 1, 1)


def _txt(s) -> str:
    return (str(s).replace("→", "->").replace("€", "EUR")
            .encode("latin-1", "replace").decode("latin-1"))


# ------------------------------------------------------------------ WORD (envuelve el compositor firmado)
def presupuesto_word(artefacto: dict, descriptor: dict, manifiesto: dict, salida: Path) -> Path:
    """Envuelve documentos/presupuesto.componer_documento (NO re-renderiza: reusa el .docx del despacho)."""
    import aqyra_documento_presupuesto as ADP
    salida = Path(salida)
    salida.parent.mkdir(parents=True, exist_ok=True)
    return ADP.componer_documento(artefacto, {"salida": salida, "fecha": _fecha(descriptor),
                                              "titulo": "PRESUPUESTO"})


# ------------------------------------------------------------------ BC3 (conecta el emisor de licitacion)
def presupuesto_bc3(artefacto: dict, descriptor: dict, manifiesto: dict, salida: Path) -> Path:
    """Conecta engines/bc3.emitir_bc3: el presupuesto sale a FIEBDC-3 (licitacion publica). Sello AAAAMMDD."""
    from aqyra_bc3 import emitir_bc3
    salida = Path(salida)
    salida.parent.mkdir(parents=True, exist_ok=True)
    fecha8 = _fecha(descriptor).replace("-", "")[:8] or None
    txt = emitir_bc3(artefacto, fecha=fecha8, charset="utf-8")
    salida.write_text(txt, encoding="utf-8", newline="")
    return salida


# ------------------------------------------------------------------ XLSX (mediciones tabular)
def presupuesto_xlsx(artefacto: dict, descriptor: dict, manifiesto: dict, salida: Path) -> Path:
    from openpyxl import Workbook
    salida = Path(salida)
    salida.parent.mkdir(parents=True, exist_ok=True)
    dt = _dt(descriptor)
    wb = Workbook()
    wb.properties.created = dt; wb.properties.modified = dt
    wb.properties.creator = manifiesto.get("generador", "aqyra-documento-export")
    ws = wb.active; ws.title = "Mediciones"
    ws.append(["Capitulo", "Codigo", "Descripcion", "Ud", "Cantidad", "Criterio de medicion",
               "Origen", "GUIDs (trazabilidad)", "Precio", "Importe"])
    for p in artefacto.get("estado_mediciones", []):
        ws.append([p.get("capitulo", ""), p.get("codigo", ""), p.get("descripcion", ""),
                   p.get("unidad", ""), round(float(p.get("cantidad", 0)), 3),
                   p.get("criterio_aplicado", ""), p.get("origen", ""),
                   ", ".join(p.get("trazabilidad", [])), round(float(p.get("precio_unitario", 0)), 2),
                   round(float(p.get("importe", 0)), 2)])
    wm = wb.create_sheet("Manifiesto")
    art = manifiesto.get("artefacto", {})
    wm.append(["Concepto", "Valor"])
    wm.append(["content_sha256", str(art.get("content_sha256"))])
    wm.append(["sello_tiempo", str(manifiesto.get("sello_tiempo"))])
    for k, v in (manifiesto.get("versiones_ancladas") or {}).items():
        wm.append([f"version.{k}", str(v)])
    wb.save(str(salida))
    return salida


# ------------------------------------------------------------------ PDF firmable (el sellado contractual)
def _h1(pdf, t):
    pdf.set_font("Helvetica", "B", 13); pdf.set_text_color(31, 78, 121)
    pdf.multi_cell(0, 8, _txt(t), new_x="LMARGIN", new_y="NEXT"); pdf.set_text_color(0, 0, 0)


def _row(pdf, cols, widths, bold=False, align=None):
    pdf.set_font("Helvetica", "B" if bold else "", 8.5)
    align = align or ["L"] * len(cols)
    for i, (c, w) in enumerate(zip(cols, widths)):
        last = i == len(cols) - 1
        pdf.cell(w, 5, _txt(c), border=1, align=align[i],
                 new_x="LMARGIN" if last else "RIGHT", new_y="NEXT" if last else "TOP")


def presupuesto_pdf(artefacto: dict, descriptor: dict, manifiesto: dict, salida: Path) -> Path:
    from fpdf import FPDF
    salida = Path(salida)
    salida.parent.mkdir(parents=True, exist_ok=True)
    resumen = artefacto.get("resumen", {})
    moneda = resumen.get("moneda", "EUR")
    em = {p["codigo"]: p for p in artefacto.get("estado_mediciones", [])}

    pdf = FPDF(format="A4"); pdf.set_auto_page_break(True, 18); pdf.set_margins(20, 18, 20)
    pdf.set_creation_date(_dt(descriptor)); pdf.set_producer("aqyra-documento-export")
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 22); pdf.set_text_color(31, 78, 121)
    pdf.multi_cell(0, 10, _txt("PRESUPUESTO"), new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 11); pdf.set_text_color(89, 89, 89)
    pdf.multi_cell(0, 6, _txt(artefacto.get("proyecto", "")), new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "B", 13); pdf.set_text_color(46, 125, 50)
    pdf.multi_cell(0, 8, _txt(f"PRESUPUESTO DE EJECUCION POR CONTRATA (PEC): "
                              f"{F.fmt_num(resumen.get('PEC', 0), 2)} {moneda}"),
                   new_x="LMARGIN", new_y="NEXT")
    pdf.set_text_color(0, 0, 0); pdf.ln(2)

    # 1 · estado de mediciones + JUSTIFICACION (criterio + origen + GUIDs = trazabilidad al modelo)
    _h1(pdf, "1. Estado de mediciones y justificacion (medicion trazada al modelo)")
    for cap in resumen.get("capitulos", []):
        pdf.set_font("Helvetica", "B", 10); pdf.set_text_color(31, 78, 121)
        pdf.multi_cell(0, 6, _txt(f"Capitulo {cap.get('codigo','')} - {cap.get('descripcion','')}"),
                       new_x="LMARGIN", new_y="NEXT"); pdf.set_text_color(0, 0, 0)
        _row(pdf, ["Codigo", "Descripcion", "Ud", "Cantidad", "Precio", "Importe"],
             [20, 58, 10, 20, 22, 22], bold=True, align=["L", "L", "L", "R", "R", "R"])
        for cod in cap.get("partidas", []):
            p = em.get(cod)
            if not p:
                continue
            _row(pdf, [cod, p.get("descripcion", "")[:44], p.get("unidad", ""),
                       F.fmt_num(p.get("cantidad", 0), 3), F.fmt_num(p.get("precio_unitario", 0), 2),
                       F.fmt_num(p.get("importe", 0), 2)],
                 [20, 58, 10, 20, 22, 22], align=["L", "L", "L", "R", "R", "R"])
            pdf.set_font("Helvetica", "I", 7.5); pdf.set_text_color(90, 90, 90)
            just = f"    Criterio: {p.get('criterio_aplicado','-')} | Origen: {p.get('origen','-')}"
            if p.get("trazabilidad"):
                just += f" | GUIDs: {', '.join(p['trazabilidad'])}"
            pdf.multi_cell(0, 4, _txt(just), new_x="LMARGIN", new_y="NEXT")
            # Slice A (D-RB-1): TEXTO ampliado de la unidad de obra (especificacion del banco), cuando
            # el salida-presupuesto lo trae (forward-open: sin `texto`, no se imprime). El ~T del BC3 ya
            # lo transporta; aqui es la evidencia contractual en el PDF sellado.
            if p.get("texto"):
                pdf.set_font("Helvetica", "", 8)
                pdf.multi_cell(0, 4, _txt(f"    {p['texto']}"), new_x="LMARGIN", new_y="NEXT")
            pdf.set_text_color(0, 0, 0)
        _row(pdf, [f"PARCIAL {cap.get('codigo','')}", "", "", "", "", F.fmt_num(cap.get("importe", 0), 2)],
             [20, 58, 10, 20, 22, 22], bold=True, align=["L", "L", "L", "R", "R", "R"])
        pdf.ln(2)

    # 2 · cuadro de precios n1 (en letra)
    pdf.add_page(); _h1(pdf, "2. Cuadro de precios n. 1 (precio unitario en letra)")
    _row(pdf, ["Codigo", "Descripcion", "Ud", "Precio", "Precio en letra"],
         [20, 44, 10, 20, 58], bold=True, align=["L", "L", "L", "R", "L"])
    for c in artefacto.get("cuadro_precios_1", []):
        _row(pdf, [c.get("codigo", ""), c.get("descripcion", "")[:34], c.get("unidad", ""),
                   F.fmt_num(c.get("precio", 0), 2), c.get("precio_en_letra", "")[:44]],
             [20, 44, 10, 20, 58], align=["L", "L", "L", "R", "L"])
    pdf.ln(2)

    # 3 · cuadro de precios n2 (descomposicion)
    _h1(pdf, "3. Cuadro de precios n. 2 (descomposicion del precio unitario)")
    for c in artefacto.get("cuadro_precios_2", []):
        pdf.set_font("Helvetica", "B", 9)
        pdf.multi_cell(0, 5, _txt(f"{c.get('codigo','')} ({c.get('unidad','')}) - Precio: "
                                  f"{F.fmt_num(c.get('precio_total', 0), 2)} {moneda}"),
                       new_x="LMARGIN", new_y="NEXT")
        _row(pdf, ["Tipo", "Descripcion", "Rend.", "Precio", "Subtotal"],
             [26, 58, 18, 20, 22], bold=True, align=["L", "L", "R", "R", "R"])
        for comp in c.get("componentes", []):
            _row(pdf, [_TIPO.get(comp.get("tipo"), comp.get("tipo", "")), comp.get("descripcion", "")[:40],
                       F.fmt_num(comp.get("rendimiento", 0), 3), F.fmt_num(comp.get("precio", 0), 2),
                       F.fmt_num(comp.get("subtotal", 0), 2)],
                 [26, 58, 18, 20, 22], align=["L", "L", "R", "R", "R"])
        _row(pdf, ["Costes indirectos", "", "", "", F.fmt_num(c.get("costes_indirectos", 0), 2)],
             [26, 58, 18, 20, 22], align=["L", "L", "R", "R", "R"])
        pdf.ln(1)

    # 4 · resumen PEM -> PEC
    _h1(pdf, "4. Resumen del presupuesto")
    for cap in resumen.get("capitulos", []):
        _row(pdf, [cap.get("codigo", ""), cap.get("descripcion", "")[:70], F.fmt_num(cap.get("importe", 0), 2)],
             [24, 96, 30], align=["L", "L", "R"])
    pdf.ln(1)
    cadena = [("Presupuesto de Ejecucion Material (PEM)", resumen.get("PEM", 0)),
              (f"Gastos generales ({F.fmt_num(resumen.get('gg_pct', 0) * 100, 0)} %)", resumen.get("gg", 0)),
              (f"Beneficio industrial ({F.fmt_num(resumen.get('bi_pct', 0) * 100, 0)} %)", resumen.get("bi", 0)),
              ("Base imponible", resumen.get("base", 0)),
              (f"IVA ({F.fmt_num(resumen.get('iva_pct', 0) * 100, 0)} %)", resumen.get("iva", 0)),
              ("Presupuesto de Ejecucion por Contrata (PEC)", resumen.get("PEC", 0))]
    for etq, val in cadena:
        _row(pdf, [etq, F.fmt_num(val, 2)], [120, 30], align=["L", "R"])

    # 5 · manifiesto de procedencia embebido
    pdf.ln(3); _h1(pdf, "5. Manifiesto de procedencia")
    art = manifiesto.get("artefacto", {})
    pdf.set_font("Helvetica", "", 8.5)
    lineas = [f"Generador: {manifiesto.get('generador')} {manifiesto.get('version_generador')}",
              f"Sello: {manifiesto.get('sello_tiempo')}  |  Artefacto: {art.get('tipo')} - {art.get('id')}",
              f"content_sha256: {art.get('content_sha256')}"]
    for k, v in (manifiesto.get("versiones_ancladas") or {}).items():
        lineas.append(f"Version {k}: {v}")
    for ln in lineas:
        pdf.multi_cell(0, 4.5, _txt(ln), new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "I", 8); pdf.set_text_color(176, 0, 0)
    pdf.multi_cell(0, 4.5, _txt("Documento contractual firmable. Integridad verificada en el gate "
                                "(Llave 1); procedencia firmada por JM (GPG, Llave 2). Preparado para la "
                                "firma cualificada/PAdES del destinatario."),
                   new_x="LMARGIN", new_y="NEXT")
    pdf.output(str(salida))
    return salida


# registro de formatos del consumidor (lo consume el nucleo export.py)
FORMATOS = {
    "word": ("Presupuesto.docx", presupuesto_word),
    "pdf":  ("Presupuesto-firmable.pdf", presupuesto_pdf),
    "bc3":  ("Presupuesto.bc3", presupuesto_bc3),
    "xlsx": ("Mediciones.xlsx", presupuesto_xlsx),
}
